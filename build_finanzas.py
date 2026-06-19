"""
Convierte el Excel de tesorería en finanzas.json (lo lee la página).

Uso:
    python build_finanzas.py "ruta/al/tesoreria_apoderados.xlsx"

Si no se pasa ruta, busca tesoreria_apoderados.xlsx en la misma carpeta.
Luego: git add finanzas.json && git commit && git push
"""
import sys, json, datetime
from pathlib import Path
import openpyxl

ruta = sys.argv[1] if len(sys.argv) > 1 else "tesoreria_apoderados.xlsx"

def num(v):
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (ValueError, TypeError):
        return 0

def txt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d-%m-%Y")
    return str(v).strip()

wb = openpyxl.load_workbook(ruta, data_only=True)

def hoja(nombre):
    return wb[nombre] if nombre in wb.sheetnames else None

# ---------- Resumen General ----------
resumen = {"cuotas_esperado": 0, "cuotas_recaudado": 0, "cuotas_diferencia": 0,
           "saldo_a_favor": 0, "gastos": 0, "saldo_disponible": 0}
ws = hoja("Resumen General")
if ws:
    for r in range(1, ws.max_row + 1):
        etq = txt(ws.cell(r, 1).value).lower()
        if etq.startswith("cuotas mensuales"):
            resumen["cuotas_esperado"] = num(ws.cell(r, 2).value)
            resumen["cuotas_recaudado"] = num(ws.cell(r, 3).value)
            resumen["cuotas_diferencia"] = num(ws.cell(r, 4).value)
        elif etq.startswith("saldo a favor"):
            resumen["saldo_a_favor"] = num(ws.cell(r, 3).value) or num(ws.cell(r, 2).value)
        elif "gastos" in etq:
            resumen["gastos"] = num(ws.cell(r, 3).value)
        elif etq.startswith("saldo disponible"):
            resumen["saldo_disponible"] = num(ws.cell(r, 3).value)

# Nota: las hojas "Cuotas Mensuales" y "Kit Semana Santa" NO se exportan,
# porque contienen nombres de familias. Solo se publican totales agregados
# (que ya vienen en "Resumen General") y el detalle de gastos por actividad.

# ---------- Gastos Actividades ----------
gastos = {"filas": [], "total": 0}
ws = hoja("Gastos Actividades")
if ws:
    for r in range(3, ws.max_row + 1):
        f = txt(ws.cell(r, 1).value)
        act = txt(ws.cell(r, 2).value)
        if f.upper().startswith("TOTAL") or act.upper().startswith("TOTAL"):
            gastos["total"] = num(ws.cell(r, 6).value)
            continue
        if not (f or act):
            continue
        gastos["filas"].append({
            "fecha": f,
            "actividad": act,
            "descripcion": txt(ws.cell(r, 3).value),
            "monto": num(ws.cell(r, 6).value),
        })
    if not gastos["total"]:
        gastos["total"] = sum(g["monto"] for g in gastos["filas"])

data = {
    "generado": datetime.date.today().isoformat(),
    "titulo": "Tesorería Apoderados 2026",
    "resumen": resumen,
    "gastos": gastos,
}

out = Path(__file__).parent / "finanzas.json"
out.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"OK -> {out}")
print(f"  resumen general + {len(gastos['filas'])} gastos (sin familias, sin kit)")
